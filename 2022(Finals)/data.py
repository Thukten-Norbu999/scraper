from openpyxl import load_workbook


def data(sheet, index_list):
    wb = load_workbook(sheet)
    sheet = wb.active
    #sheet = wb.get_sheet_by_name('Sheet1')
    max_row = sheet.max_row
    
    index, dob = (sheet.cell(row=2, column=1).value), sheet.cell(row=2, column=2).value
    new = str(dob).split(' ')[0].split('-')
    dob = f'{new[2]}/{new[1]}/{new[0]}'
    return f'0{str(index)[:11]}', str(dob), new
print(data('./BHSEC/Uploads/rq.xlsx', []))