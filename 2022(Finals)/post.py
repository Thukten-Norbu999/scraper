from openpyxl import load_workbook
import os
import openpyxl
def give(sheet, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in data:
        ws.append(i)
    wb.save(f'./BHSEC/Outputs/{sheet}')
give('output.xlsx',(("THukten","KINga","Lunda"),("THukten","KINga","Lunda"),("THukten","KINga","Lunda")))
'''
wb = openpyxl.Workbook()
ws = wb.active

data = ( 
("Product","Cost Price","Selling Price"), 
("earpod",90, 50), 
("laptop", 3000, 8200), 
("smartphone", 5100, 7200) 
)

for i in data:
   ws.append(i)

wb.save('test.xlsx')
'''
