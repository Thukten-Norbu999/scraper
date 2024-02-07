#SELENIUM
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By

import time
import os

#
from openpyxl import load_workbook


class ScraperB:

    def __init__(self, individual):
        self.individual = individual
        self.PATH = "chromedriver.exe"
        self.service = Service(executable_path=self.PATH)
        self.driver = webdriver.Chrome(service=self.service)

    def scrape(self, INDEX, DOB):
        self.driver.get("https://class12result.cloud/result2021")
        self.index = self.driver.find_element(By.XPATH, "//*[@id='indexno']")
        self.index.send_keys(INDEX)
        self.index.send_keys(Keys.RETURN)

        self.dob = self.driver.find_element(By.XPATH, "//*[@id='dob']")
        #dob.clear()
        self.dob.send_keys(DOB)
        self.dob.send_keys(Keys.RETURN)

        c_index = self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[1]/td').text
        c_Math, c_Bio = self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[6]/th[1]').text, self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[9]/th[1]').text

        if c_index == INDEX:
        #MARKS
            cid = self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[2]/td').text
            eng = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[4]/th[2]').text)
            dzo = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[5]/th[2]').text)


            if c_Math == "MATHEMATICS" and c_Bio == 'BIOLOGY':
                math = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[6]/th[2]').text)
                #Sci Sub #
                phy = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[7]/th[2]').text)
                chem = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[8]/th[2]').text)
                bio = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[9]/th[2]').text)
                perc = sum(sorted([eng, dzo, math, phy, chem, bio])[2:])/4
                entry = [INDEX, DOB, cid, eng, dzo, math, phy, chem, bio, perc]
                return entry
            else:
                if c_Math == "MATHEMATICS":
                    math = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[6]/th[2]').text)
                    phy = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[7]/th[2]').text)
                    chem = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[8]/th[2]').text)
                    bio = "NA"
                    perc = sum(sorted([eng, dzo, phy, chem, math])[1:])/4
                    
                    entry = [INDEX, DOB, cid, eng, dzo, math, phy, chem, bio, perc]
                    return entry

                else:
                    math = "NA"
                    phy = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[6]/th[2]').text)
                    chem = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[7]/th[2]').text)
                    bio = float(self.driver.find_element(By.XPATH, '//*[@id="myTable"]/tbody/tr[8]/th[2]').text)
                    perc = sum(sorted([eng, dzo, phy, chem, bio])[1:])/4

                    entry = [INDEX, DOB, cid, eng, dzo, math, phy, chem, bio, perc]
                    return entry

        self.driver.quit()
    
    def getdata(self, file_name):
        wb = load_workbook(file_name)
        sheet = wb.active
        sheet = wb.get_sheet_by_name('Sheet1')
        info_list = []

        max_row, max_col = sheet.max_row, sheet.max_column

        for i in range(1, max_col-1):
            for col in sheet.iter_cols(max_col=1,min_row=2):
                for cell in col:
                    info_list.append(cell.value)
        



marks = ScraperB(False)

marks.getdata('file.xlsx')
