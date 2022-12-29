from openpyxl import load_workbook


def data(sheet, index_list):
    wb = load_workbook(sheet)
    sheet = wb.active
    sheet = wb.get_sheet_by_name('Sheet1')

    for col in sheet.iter_cols(max_col=1,min_row=2):
        for cell in col:
            index_list.append(cell.value)
        
